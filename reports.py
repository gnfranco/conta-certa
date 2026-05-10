from __future__ import annotations

from datetime import datetime
from html import escape
from io import BytesIO
from typing import Any

import pandas as pd

from calculos import money


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _format_date_br(value: Any) -> str:
    if not value:
        return ""

    text = str(value)[:10]

    try:
        year, month, day = text.split("-")
        return f"{day}/{month}/{year}"
    except Exception:
        return str(value)


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_paragraph(value: Any, style):
    from reportlab.platypus import Paragraph

    return Paragraph(escape(_safe_text(value)), style)


def _titulos_rows(resultado: dict[str, Any]) -> list[dict[str, Any]]:
    titulos = resultado.get("titulos", resultado.get("dividas", []))
    rows: list[dict[str, Any]] = []

    for t in titulos:
        rows.append(
            {
                "Ref.": t.get("titulo_ref") or t.get("public_ref") or "",
                "Lote": t.get("lote_ref") or "",
                "Devedor": t.get("devedor") or "",
                "Grupo": t.get("grupo") or "",
                "Tipo": t.get("tipo") or "",
                "Competência": t.get("competencia") or "",
                "Descrição": t.get("descricao") or "",
                "Vencimento": _format_date_br(t.get("vencimento")),
                "Valor original": _as_float(t.get("valor_original")),
                "Recebido": _as_float(t.get("total_recebido")),
                "Principal aberto": _as_float(t.get("principal_aberto_estimado")),
                "Encargos": _as_float(t.get("encargos")),
                "Saldo atualizado": _as_float(t.get("saldo_atualizado")),
                "Situação": t.get("situacao_financeira") or "",
                "Status adm.": t.get("status_administrativo") or "",
                "Índices provisórios": t.get("competencias_provisorias") or "",
                "Índices faltando": t.get("competencias_faltando") or "",
            }
        )

    return rows


def _baixas_rows(resultado: dict[str, Any]) -> list[dict[str, Any]]:
    baixas = resultado.get("baixas", resultado.get("alocacoes", []))
    rows: list[dict[str, Any]] = []

    for b in baixas:
        rows.append(
            {
                "Recebimento": b.get("pagamento_ref") or "",
                "Data": _format_date_br(b.get("data_pagamento")),
                "Devedor": b.get("devedor") or "",
                "Grupo": b.get("grupo") or "",
                "Título": b.get("titulo_ref") or b.get("divida_ref") or "",
                "Lote": b.get("lote_ref") or "",
                "Histórico": b.get("titulo") or b.get("divida") or "",
                "Tipo": b.get("tipo_alocacao") or "Baixa",
                "Valor alocado": _as_float(b.get("valor_alocado")),
            }
        )

    return rows


def _indices_rows(taxas_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for t in taxas_rows:
        rows.append(
            {
                "Competência": t.get("competencia") or "",
                "IPCA (%)": t.get("ipca_pct"),
                "Taxa Legal (%)": t.get("taxa_legal_pct"),
                "Fonte": t.get("fonte") or "",
                "Status": t.get("status") or "",
                "Atualizado em": t.get("updated_at") or "",
            }
        )

    return rows


def _resumo_rows(resultado: dict[str, Any]) -> list[dict[str, Any]]:
    resumo = resultado["resumo"]

    return [
        {"Campo": "Data-base", "Valor": _format_date_br(resumo.get("data_base"))},
        {"Campo": "Total original vencido", "Valor": _as_float(resumo.get("total_original_vencido"))},
        {"Campo": "Recebimentos considerados", "Valor": _as_float(resumo.get("pagamentos_considerados"))},
        {"Campo": "Principal aberto", "Valor": _as_float(resumo.get("principal_aberto_estimado"))},
        {"Campo": "Encargos IPCA + Taxa Legal", "Valor": _as_float(resumo.get("encargos"))},
        {"Campo": "Total atualizado", "Valor": _as_float(resumo.get("total_atualizado"))},
        {"Campo": "Créditos/excedentes não alocados", "Valor": _as_float(resumo.get("creditos_excedentes"))},
        {"Campo": "Títulos vencidos", "Valor": int(resumo.get("titulos_vencidos", 0) or 0)},
        {"Campo": "Títulos parciais", "Valor": int(resumo.get("titulos_parciais", 0) or 0)},
        {"Campo": "Títulos quitados", "Valor": int(resumo.get("titulos_quitados", 0) or 0)},
        {"Campo": "Competências provisórias", "Valor": resumo.get("competencias_provisorias") or ""},
        {"Campo": "Competências faltando", "Valor": resumo.get("competencias_faltando") or ""},
        {"Campo": "Emitido em", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
    ]


def _apply_excel_style(writer: pd.ExcelWriter) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    money_columns = {
        "Valor original",
        "Recebido",
        "Principal aberto",
        "Encargos",
        "Saldo atualizado",
        "Valor alocado",
    }

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        header_by_col = {cell.column: str(cell.value or "") for cell in ws[1]}

        for col_cells in ws.columns:
            col_index = col_cells[0].column
            col_letter = get_column_letter(col_index)
            header = header_by_col.get(col_index, "")
            max_len = 10

            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))

                if header in money_columns and cell.row > 1:
                    cell.number_format = '"R$" #,##0.00'

                if header in {"IPCA (%)", "Taxa Legal (%)"} and cell.row > 1:
                    cell.number_format = '0.0000'

                cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def gerar_excel(resultado: dict[str, Any], taxas_rows: list[dict[str, Any]]) -> bytes:
    output = BytesIO()

    df_resumo = pd.DataFrame(_resumo_rows(resultado))
    df_titulos = pd.DataFrame(_titulos_rows(resultado))
    df_baixas = pd.DataFrame(_baixas_rows(resultado))
    df_indices = pd.DataFrame(_indices_rows(taxas_rows))

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_titulos.to_excel(writer, sheet_name="Titulos", index=False)
        df_baixas.to_excel(writer, sheet_name="Baixas", index=False)
        df_indices.to_excel(writer, sheet_name="Indices", index=False)

        _apply_excel_style(writer)

    return output.getvalue()


def gerar_pdf(
    resultado: dict[str, Any],
    titulo: str = "Demonstrativo Conta Certa",
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.1 * cm,
        rightMargin=1.1 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=7, leading=8))
    styles.add(ParagraphStyle(name="SmallBold", parent=styles["Small"], fontName="Helvetica-Bold"))

    story = []
    resumo = resultado["resumo"]

    story.append(Paragraph(escape(titulo), styles["Title"]))
    story.append(
        Paragraph(
            f"Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} · "
            f"Data-base {_format_date_br(resumo.get('data_base'))}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    cards = [
        ["Principal aberto", money(_as_float(resumo.get("principal_aberto_estimado")))],
        ["Encargos", money(_as_float(resumo.get("encargos")))],
        ["Total atualizado", money(_as_float(resumo.get("total_atualizado")))],
        ["Recebimentos", money(_as_float(resumo.get("pagamentos_considerados")))],
        ["Créditos/excedentes", money(_as_float(resumo.get("creditos_excedentes")))],
    ]

    card_table = Table(cards, colWidths=[4.0 * cm, 4.0 * cm])
    card_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(card_table)
    story.append(Spacer(1, 0.25 * cm))

    status_rows = [
        [
            "Títulos vencidos",
            str(int(resumo.get("titulos_vencidos", 0) or 0)),
            "Títulos parciais",
            str(int(resumo.get("titulos_parciais", 0) or 0)),
            "Títulos quitados",
            str(int(resumo.get("titulos_quitados", 0) or 0)),
        ]
    ]

    status_table = Table(status_rows, colWidths=[3.2 * cm, 1.3 * cm, 3.2 * cm, 1.3 * cm, 3.2 * cm, 1.3 * cm])
    status_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("ALIGN", (3, 0), (3, 0), "CENTER"),
                ("ALIGN", (5, 0), (5, 0), "CENTER"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    story.append(status_table)
    story.append(Spacer(1, 0.25 * cm))

    if resumo.get("competencias_provisorias"):
        story.append(Paragraph("Índices provisórios usados: " + escape(str(resumo["competencias_provisorias"])), styles["Small"]))

    if resumo.get("competencias_faltando"):
        story.append(Paragraph("Índices faltando: " + escape(str(resumo["competencias_faltando"])), styles["Small"]))

    story.append(Spacer(1, 0.35 * cm))
    story.append(Paragraph("Títulos atualizados", styles["Heading2"]))

    titulos = _titulos_rows(resultado)
    titulo_rows = [["Ref.", "Lote", "Devedor", "Grupo", "Compet.", "Venc.", "Principal", "Encargos", "Saldo", "Situação"]]

    for t in titulos:
        titulo_rows.append(
            [
                _safe_paragraph(t["Ref."], styles["Small"]),
                _safe_paragraph(t["Lote"], styles["Small"]),
                _safe_paragraph(t["Devedor"][:25], styles["Small"]),
                _safe_paragraph(t["Grupo"][:18], styles["Small"]),
                _safe_paragraph(t["Competência"], styles["Small"]),
                _safe_paragraph(t["Vencimento"], styles["Small"]),
                money(_as_float(t["Principal aberto"])),
                money(_as_float(t["Encargos"])),
                money(_as_float(t["Saldo atualizado"])),
                _safe_paragraph(t["Situação"], styles["Small"]),
            ]
        )

    if len(titulo_rows) > 1:
        t_titulos = Table(
            titulo_rows,
            repeatRows=1,
            colWidths=[2.7 * cm, 2.6 * cm, 3.3 * cm, 2.5 * cm, 1.8 * cm, 1.9 * cm, 2.4 * cm, 2.1 * cm, 2.4 * cm, 2.5 * cm],
        )
        t_titulos.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (6, 1), (8, -1), "RIGHT"),
                    ("PADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(t_titulos)
    else:
        story.append(Paragraph("Nenhum título considerado.", styles["Normal"]))

    story.append(PageBreak())
    story.append(Paragraph("Baixas, recebimentos e excedentes", styles["Heading2"]))

    baixas = _baixas_rows(resultado)
    baixa_rows = [["Recebimento", "Data", "Devedor", "Grupo", "Título", "Lote", "Histórico", "Tipo", "Valor"]]

    for b in baixas:
        baixa_rows.append(
            [
                _safe_paragraph(b["Recebimento"], styles["Small"]),
                _safe_paragraph(b["Data"], styles["Small"]),
                _safe_paragraph(b["Devedor"][:25], styles["Small"]),
                _safe_paragraph(b["Grupo"][:18], styles["Small"]),
                _safe_paragraph(b["Título"], styles["Small"]),
                _safe_paragraph(b["Lote"], styles["Small"]),
                _safe_paragraph(b["Histórico"][:32], styles["Small"]),
                _safe_paragraph(b["Tipo"], styles["Small"]),
                money(_as_float(b["Valor alocado"])),
            ]
        )

    if len(baixa_rows) > 1:
        t_baixas = Table(
            baixa_rows,
            repeatRows=1,
            colWidths=[3.0 * cm, 1.8 * cm, 3.3 * cm, 2.6 * cm, 2.8 * cm, 2.6 * cm, 4.3 * cm, 2.0 * cm, 2.2 * cm],
        )
        t_baixas.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (8, 1), (8, -1), "RIGHT"),
                    ("PADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(t_baixas)
    else:
        story.append(Paragraph("Nenhuma baixa ou recebimento considerado.", styles["Normal"]))

    story.append(Spacer(1, 0.45 * cm))
    story.append(Paragraph("Critérios do demonstrativo", styles["Heading2"]))
    story.append(
        Paragraph(
            "Este demonstrativo considera os títulos vencidos até a data-base informada, "
            "os recebimentos registrados até a mesma data e a atualização por IPCA + Taxa Legal. "
            "Quando uma competência ainda não possui índice oficial cadastrado, o sistema pode usar "
            "índices provisórios, conforme configuração local.",
            styles["Small"],
        )
    )

    doc.build(story)
    return buffer.getvalue()
